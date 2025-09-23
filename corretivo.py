#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Corretivo: reprocessa tickets do Freshdesk a partir de um CSV de erros,
baixa anexos por ticket (ticket e conversas) para pastas por ID, e
atualiza mapeamentos Octa ID (contatos/organizações) no MySQL usando
uma planilha local (octadesk_data.xlsx por padrão).

Requisitos-chave implementados:
- Lê CSV (--errors-file) e extrai IDs de tickets a reprocessar.
- Para cada ticket, cria pasta <download-dir>/<ticket_id> e baixa todos os anexos
  do ticket e de suas conversas, respeitando --min-attach-kb e evitando duplicar
  com marcador .state/ticket_<id>.done sob a pasta base.
- Usa planilha de mapeamento (--mapping-file) para vincular Contatos (id_contato→Octa ID)
  e Organizações (org_id→Octa ID) e atualiza no MySQL:
    contacts.octa_contact_id WHERE contacts.freshdesk_id = id_contato
    companies.octa_company_id WHERE companies.fresh_company_id = org_id
  (tabelas e colunas podem ser trocadas por .env/CLI)
- Apenas API Freshdesk é consumida (conversas/anexos). Octadesk NÃO é chamado.
- Tratamento de rate limit 429 com backoff exponencial e Retry-After.
- Logs em console e arquivo (--log-file) com níveis [INFO]/[WARNING]/[ERROR].
- Parâmetros CLI conforme especificado.

Dependências sugeridas: requests, python-dotenv (opcional), openpyxl, mysql-connector-python


Comando: python corretivo.py --errors-file .\errors_<mes>.csv --download-dir "C:\anexos_freshdesk_tickets" --min-attach-kb 0 --batch-size 400 --mapping-file .\data\octadesk_data.xlsx --log-file corretivo_<mes>.log
"""
from __future__ import annotations

import argparse
import csv
import logging
import os
import re
import sys
import time
import shutil  # <-- ADICIONADO PARA MOVER ARQUIVOS
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

import requests
from requests.auth import HTTPBasicAuth

try:
    import mysql.connector as mysql
except Exception as e:
    print("[ERROR] mysql-connector-python não está instalado: pip install mysql-connector-python", file=sys.stderr)
    raise

try:
    from openpyxl import load_workbook
except Exception:
    print("[ERROR] openpyxl não está instalado: pip install openpyxl", file=sys.stderr)
    raise

# Logger global para evitar WARNING:root e unificar saída
LOGGER = logging.getLogger("corretivo")

# -------- Util: .env loader simples (sem dependência externa) --------
def load_dotenv(path: str = ".env") -> None:
    if not os.path.isfile(path):
        return
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip().strip('"').strip("'")
                if k and (k not in os.environ):
                    os.environ[k] = v
    except Exception as e:
        LOGGER.warning("Falha ao ler .env: %s", e)

# -------- Configuração de logging --------
class ConsoleFileLogger:
    def __init__(self, logfile: Optional[str]):
        global LOGGER
        LOGGER.setLevel(logging.INFO)
        fmt = logging.Formatter("[%(levelname)s] %(asctime)s - %(message)s")
        if not LOGGER.handlers:
            sh = logging.StreamHandler(sys.stdout)
            sh.setFormatter(fmt)
            LOGGER.addHandler(sh)
        if logfile:
            # Garante que o manipulador de arquivo seja removido se já existir para evitar duplicação
            for handler in LOGGER.handlers[:]:
                if isinstance(handler, logging.FileHandler):
                    LOGGER.removeHandler(handler)
            fh = logging.FileHandler(logfile, encoding="utf-8")
            fh.setFormatter(fmt)
            LOGGER.addHandler(fh)
        LOGGER.propagate = False

    def get(self):
        return LOGGER

# -------- Config e argumentos --------
@dataclass
class EnvConfig:
    fd_subdomain: str
    fd_api_key: str
    mysql_host: str
    mysql_port: int
    mysql_user: str
    mysql_password: str
    mysql_db: str
    table_contacts: str
    table_companies: str
    contacts_octa_id_field: str
    companies_octa_id_field: str
    db_col_contact_pk: str
    db_col_company_pk: str

def read_env(env_path: str) -> EnvConfig:
    load_dotenv(env_path)
    return EnvConfig(
        fd_subdomain=os.getenv("FRESHDESK_SUBDOMAIN", ""),
        fd_api_key=os.getenv("FRESHDESK_API_KEY", ""),
        mysql_host=os.getenv("MYSQL_HOST", "127.0.0.1"),
        mysql_port=int(os.getenv("MYSQL_PORT", "3306")),
        mysql_user=os.getenv("MYSQL_USER", "root"),
        mysql_password=os.getenv("MYSQL_PASSWORD", ""),
        mysql_db=os.getenv("MYSQL_DATABASE", ""),
        table_contacts=os.getenv("TABLE_CONTACTS", "contacts"),
        table_companies=os.getenv("TABLE_COMPANIES", "companies"),
        contacts_octa_id_field=os.getenv("CONTACTS_OCTA_ID_FIELD", "octa_contact_id"),
        companies_octa_id_field=os.getenv("COMPANIES_OCTA_ID_FIELD", "octa_company_id"),
        db_col_contact_pk=os.getenv("DB_COL_CONTACT_PK", "freshdesk_id"),
        db_col_company_pk=os.getenv("DB_COL_COMPANY_PK", "fresh_company_id"),
    )

# -------- Freshdesk helpers --------
def fd_base(subdomain: str) -> str:
    sd = (subdomain or "").strip().rstrip("/")
    if sd and "." not in sd:
        sd = f"{sd}.freshdesk.com"
    if not sd.startswith("http"):
        sd = "https://" + sd
    return sd

def fd_get(subdomain: str, api_key: str, path: str, params: Optional[Dict[str, str]] = None,
           max_retries: int = 5, session: Optional[requests.Session] = None) -> requests.Response:
    url = f"{fd_base(subdomain)}/api/v2{path}"
    sess = session or requests.Session()
    attempt = 0
    while True:
        try:
            r = sess.get(url, params=params or {}, headers={"Accept": "application/json"},
                         auth=HTTPBasicAuth(api_key, "X"), timeout=120)
            if r.status_code == 429:
                retry_after = r.headers.get("Retry-After")
                wait = float(retry_after) if retry_after and retry_after.isdigit() else (2 ** attempt)
                wait = min(wait, 60.0)
                LOGGER.warning("[429] Rate limit Freshdesk. Aguardando %.1fs...", wait)
                time.sleep(wait)
                attempt += 1
                if attempt > max_retries:
                    r.raise_for_status()
                continue
            r.raise_for_status()
            return r
        except requests.RequestException as e:
            if attempt < max_retries:
                wait = min(2 ** attempt, 30)
                LOGGER.warning("[WARN] Falha de rede em %s: %s. Retentando em %ss...", path, e, wait)
                time.sleep(wait)
                attempt += 1
                continue
            raise

def fd_get_ticket_full(subdomain: str, api_key: str, ticket_id: int, session: Optional[requests.Session]=None) -> dict:
    r = fd_get(subdomain, api_key, f"/tickets/{ticket_id}", params={"include": "conversations"}, session=session)
    return r.json()

# -------- Download de anexos --------
ATT_NAME_SIG_RE = re.compile(r"(logo|assinatura|signature|rodape|footer|image0+\d|facebook|instagram|linkedin|twitter|whatsapp)", re.I)

def safe_filename(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*\x00-\x1F]', "_", name)[:180]

def ensure_dir(p: Path):
    p.mkdir(parents=True, exist_ok=True)

def download_binary(url: str, dest: Path, min_bytes: int, session: Optional[requests.Session]=None) -> Tuple[bool, int]:
    sess = session or requests.Session()
    try:
        with sess.get(url, timeout=120, stream=True) as rr:
            rr.raise_for_status()
            size = int(rr.headers.get("Content-Length", 0))
            if size < min_bytes:
                LOGGER.info("[INFO] Ignorado anexo muito pequeno (%d B): %s", size, url)
                return False, size
            ensure_dir(dest.parent)
            with open(dest, "wb") as f:
                for chunk in rr.iter_content(chunk_size=8192):
                    f.write(chunk)
            return True, size
    except requests.RequestException as e:
        LOGGER.warning("[WARNING] Falha ao baixar %s: %s", url, e)
        return False, 0

def collect_and_download_attachments(ticket: dict, ticket_dir: Path, min_kb: int,
                                     block_signature_like: bool = True,
                                     session: Optional[requests.Session]=None) -> int:
    saved = 0
    min_bytes = max(0, min_kb) * 1024

    def handle_one(name: str, url: str):
        nonlocal saved
        if not url: return
        fn = safe_filename(name or url.split("/")[-1].split("?")[0] or "attachment")
        if block_signature_like and ATT_NAME_SIG_RE.search(fn):
            LOGGER.info("[INFO] Pulado provável assinatura/logo: %s", fn)
            return
        ok, size = download_binary(url, ticket_dir / fn, min_bytes, session=session)
        if ok:
            saved += 1
            LOGGER.info("[INFO] Anexo salvo (%d B): %s", size, fn)

    for a in (ticket.get("attachments") or []):
        handle_one(a.get("name"), a.get("attachment_url"))
    for c in (ticket.get("conversations") or []):
        for a in (c.get("attachments") or []):
            handle_one(a.get("name"), a.get("attachment_url"))
    return saved

# -------- Planilha de mapeamento --------
@dataclass
class Mapping:
    contact_by_fd_id: Dict[int, str]
    org_by_fd_id: Dict[int, str]

def load_mapping(xlsx_path: str) -> Mapping:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    contacts_map: Dict[int, str] = {}
    if "Contatos" in wb.sheetnames:
        ws = wb["Contatos"]
        headers = {str(cell.value).strip().lower(): idx for idx, cell in enumerate(ws[1])}
        col_id = next((headers[h] for h in headers if h in ("id_contato", "freshdesk_id")), 0)
        col_octa = next((headers[h] for h in headers if h in ("octa id", "octa_id")), 1)
        for row in ws.iter_rows(min_row=2):
            try:
                fdid = int(str(row[col_id].value).strip())
                octa = str(row[col_octa].value).strip()
                if fdid and octa: contacts_map[fdid] = octa
            except (ValueError, TypeError): continue
    
    org_map: Dict[int, str] = {}
    ws_name = next((s for s in wb.sheetnames if s.lower() in ("organizações", "organizacoes")), None)
    if ws_name:
        ws = wb[ws_name]
        headers = {str(cell.value).strip().lower(): idx for idx, cell in enumerate(ws[1])}
        col_id = next((headers[h] for h in headers if h in ("org_id", "fresh_company_id")), 0)
        col_octa = next((headers[h] for h in headers if h in ("octa id", "octa_id")), 1)
        for row in ws.iter_rows(min_row=2):
            try:
                fdid = int(str(row[col_id].value).strip())
                octa = str(row[col_octa].value).strip()
                if fdid and octa: org_map[fdid] = octa
            except (ValueError, TypeError): continue

    return Mapping(contact_by_fd_id=contacts_map, org_by_fd_id=org_map)

# -------- Atualização MySQL --------
def mysql_connect(cfg: EnvConfig):
    return mysql.connect(
        host=cfg.mysql_host, port=cfg.mysql_port, user=cfg.mysql_user,
        password=cfg.mysql_password, database=cfg.mysql_db,
        charset="utf8mb4", use_unicode=True, autocommit=False
    )

def apply_mappings(conn, cfg: EnvConfig, mapping: Mapping, report: Optional[Path] = None) -> Tuple[int, int]:
    c = conn.cursor()
    contacts_updated, orgs_updated = 0, 0
    writer = None
    repw = None
    if report:
        ensure_dir(report.parent)
        repw = open(report, "w", encoding="utf-8", newline="")
        writer = csv.writer(repw)
        writer.writerow(["tipo", "freshdesk_id", "octa_id", "status"])

    sql_contacts = (f"UPDATE `{cfg.table_contacts}` SET `{cfg.contacts_octa_id_field}` = %s "
                    f"WHERE `{cfg.db_col_contact_pk}` = %s AND NOT (`{cfg.contacts_octa_id_field}` <=> %s)")
    sql_companies = (f"UPDATE `{cfg.table_companies}` SET `{cfg.companies_octa_id_field}` = %s "
                     f"WHERE `{cfg.db_col_company_pk}` = %s AND NOT (`{cfg.companies_octa_id_field}` <=> %s)")

    def _exists(table: str, col: str, fdid: int) -> bool:
        c.execute(f"SELECT 1 FROM `{table}` WHERE `{col}`=%s LIMIT 1", (fdid,))
        return bool(c.fetchone())

    for fdid, octa in mapping.contact_by_fd_id.items():
        try:
            c.execute(sql_contacts, (octa, fdid, octa))
            if c.rowcount > 0:
                contacts_updated += c.rowcount
                LOGGER.info("[INFO] Contato %s atualizado para %s.", fdid, octa)
                if writer: writer.writerow(["contato", fdid, octa, "atualizado"])
            else:
                if _exists(cfg.table_contacts, cfg.db_col_contact_pk, fdid):
                    LOGGER.info("[INFO] Contato %s já estava com o Octa ID correto.", fdid)
                    if writer: writer.writerow(["contato", fdid, octa, "ja_igual"])
                else:
                    LOGGER.warning("[WARNING] Contato não encontrado p/ update: WHERE %s=%s na tabela %s",
                                   cfg.db_col_contact_pk, fdid, cfg.table_contacts)
                    if writer: writer.writerow(["contato", fdid, octa, "inexistente"])
            conn.commit()
        except Exception as e:
            conn.rollback()
            LOGGER.error("[ERROR] Falha ao atualizar contato %s -> %s: %s", fdid, octa, e)

    for fdid, octa in mapping.org_by_fd_id.items():
        try:
            c.execute(sql_companies, (octa, fdid, octa))
            if c.rowcount > 0:
                orgs_updated += c.rowcount
                LOGGER.info("[INFO] Organização %s atualizada para %s.", fdid, octa)
                if writer: writer.writerow(["organizacao", fdid, octa, "atualizado"])
            else:
                if _exists(cfg.table_companies, cfg.db_col_company_pk, fdid):
                    LOGGER.info("[INFO] Organização %s já estava com o Octa ID correto.", fdid)
                    if writer: writer.writerow(["organizacao", fdid, octa, "ja_igual"])
                else:
                    LOGGER.warning("[WARNING] Empresa não encontrada p/ update: WHERE %s=%s na tabela %s",
                                   cfg.db_col_company_pk, fdid, cfg.table_companies)
                    if writer: writer.writerow(["organizacao", fdid, octa, "inexistente"])
            conn.commit()
        except Exception as e:
            conn.rollback()
            LOGGER.error("[ERROR] Falha ao atualizar organização %s -> %s: %s", fdid, octa, e)

    if repw:
        repw.close()
        LOGGER.info("[INFO] Relatório de reconciliação salvo em: %s", str(report))
    c.close()
    return contacts_updated, orgs_updated

# -------- Processamento --------
def parse_error_data_from_csv(path: str) -> Tuple[List[int], Set[int], Set[int]]:
    """Lê um CSV de erros e extrai IDs de tickets, contatos e empresas.
       Retorna uma tupla com: (lista de ticket_ids, conjunto de contact_ids, conjunto de company_ids)
    """
    ticket_ids: List[int] = []
    contact_ids: Set[int] = set()
    company_ids: Set[int] = set()

    probable_ticket_cols = {"ticket_id", "ticket", "id"}
    probable_contact_cols = {"contact_fresh_id", "contact_id", "id_contato"}
    probable_company_cols = {"company_fresh_id", "company_id", "org_id", "company_name"}

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return [], set(), set()

    header = [h.strip().lower() for h in rows[0]]
    idx_map = {h: i for i, h in enumerate(header)}

    ticket_col_idx = next((idx_map[h] for h in header if h in probable_ticket_cols), None)
    contact_col_idx = next((idx_map[h] for h in header if h in probable_contact_cols), None)
    company_col_idx = next((idx_map[h] for h in header if h in probable_company_cols), None)

    if ticket_col_idx is None:
        raise ValueError("Coluna de 'ticket_id' não encontrada no CSV de erros.")

    for r in rows[1:]:
        if ticket_col_idx < len(r) and r[ticket_col_idx].strip().isdigit():
            ticket_ids.append(int(r[ticket_col_idx].strip()))
        
        if contact_col_idx is not None and contact_col_idx < len(r) and r[contact_col_idx].strip().isdigit():
            contact_ids.add(int(r[contact_col_idx].strip()))

        if company_col_idx is not None and company_col_idx < len(r) and r[company_col_idx].strip().isdigit():
            company_ids.add(int(r[company_col_idx].strip()))

    seen_tickets = set()
    unique_ticket_ids = [i for i in ticket_ids if not (i in seen_tickets or seen_tickets.add(i))]

    return unique_ticket_ids, contact_ids, company_ids

def process_tickets(tickets: List[int], cfg: EnvConfig, download_dir: str,
                    min_attach_kb: int, batch_size: int, state_dir: Path,
                    block_signature_like: bool, logger: logging.Logger) -> Tuple[int, int]:
    sess = requests.Session()
    processed, saved_total = 0, 0

    for i in range(0, len(tickets), batch_size):
        batch = tickets[i:i+batch_size]
        logger.info("[INFO] Lote de %d tickets (%d-%d)", len(batch), i+1, i+len(batch))
        for tid in batch:
            marker = state_dir / f"ticket_{tid}.done"
            if marker.exists():
                logger.info("[INFO] Ticket %s já marcado como concluído. Pulando download.", tid)
                processed += 1
                continue
            try:
                full = fd_get_ticket_full(cfg.fd_subdomain, cfg.fd_api_key, tid, session=sess)
                tdir = Path(download_dir) / str(tid)
                ensure_dir(tdir)
                saved = collect_and_download_attachments(full, tdir, min_attach_kb, block_signature_like, session=sess)
                saved_total += saved
                logger.info("[INFO] Ticket %s: %d anexos salvos.", tid, saved)
                ensure_dir(state_dir)
                marker.touch()
                processed += 1
            except requests.HTTPError as he:
                logger.error("[ERROR] Falha ao buscar ticket %s (HTTP %s)", tid, he.response.status_code if he.response else "?")
            except Exception as e:
                logger.error("[ERROR] Falha inesperada ao processar ticket %s: %s", tid, e)

    return processed, saved_total

# -------- CLI --------
def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Corretivo Freshdesk: baixar anexos e atualizar mapeamentos.")
    p.add_argument("--errors-file", required=True, help="CSV com os tickets a reprocessar")
    p.add_argument("--download-dir", required=True, help="Pasta base dos anexos")
    p.add_argument("--min-attach-kb", type=int, default=5, help="Tamanho mínimo de anexo para salvar (kB)")
    p.add_argument("--batch-size", type=int, default=50, help="Quantidade de tickets por lote")
    p.add_argument("--mapping-file", default="octadesk_data.xlsx", help="Planilha com mapeamento")
    p.add_argument("--log-file", default=None, help="Caminho do arquivo de log")
    p.add_argument("--env-file", default=".env", help="Caminho do arquivo .env")
    p.add_argument("--no-attach-signature-block", action="store_true", help="Não bloquear anexos de assinatura/logo")
    p.add_argument("--report-file", default=None, help="CSV de reconciliação contato/organização vs banco")
    return p

def main():
    args = build_arg_parser().parse_args()
    logger = ConsoleFileLogger(args.log_file).get()

    cfg = read_env(args.env_file)
    if not all([cfg.fd_subdomain, cfg.fd_api_key, cfg.mysql_db]):
        logger.error("Variáveis obrigatórias ausentes no .env: FRESHDESK_SUBDOMAIN, FRESHDESK_API_KEY, MYSQL_DATABASE")
        sys.exit(2)

    try:
        ticket_ids, contact_ids_error, company_ids_error = parse_error_data_from_csv(args.errors_file)
        logger.info("Dados de erro carregados: %d tickets, %d contatos, %d organizações a verificar.",
                    len(ticket_ids), len(contact_ids_error), len(company_ids_error))
    except Exception as e:
        logger.error("Falha ao ler --errors-file: %s", e)
        sys.exit(2)

    try:
        full_mapping = load_mapping(args.mapping_file)
        logger.info("Mapeamentos totais carregados: %d contatos, %d organizações.",
                    len(full_mapping.contact_by_fd_id), len(full_mapping.org_by_fd_id))
    except Exception as e:
        logger.error("Falha ao carregar --mapping-file: %s", e)
        sys.exit(2)
        
    filtered_contact_map = {fd_id: octa_id for fd_id, octa_id in full_mapping.contact_by_fd_id.items() if fd_id in contact_ids_error}
    filtered_org_map = {fd_id: octa_id for fd_id, octa_id in full_mapping.org_by_fd_id.items() if fd_id in company_ids_error}
    filtered_mapping = Mapping(contact_by_fd_id=filtered_contact_map, org_by_fd_id=filtered_org_map)
    
    logger.info("Mapeamentos filtrados para atualização: %d contatos, %d organizações.",
                len(filtered_mapping.contact_by_fd_id), len(filtered_mapping.org_by_fd_id))

    if filtered_mapping.contact_by_fd_id or filtered_mapping.org_by_fd_id:
        try:
            conn = mysql_connect(cfg)
            report_path = Path(args.report_file) if args.report_file else None
            c_upd, o_upd = apply_mappings(conn, cfg, filtered_mapping, report=report_path)
            logger.info("Atualizações de mapeamento concluídas: contacts=%d, companies=%d", c_upd, o_upd)
            conn.close()
        except Exception as e:
            logger.error("Erro ao aplicar mapeamentos no MySQL: %s", e)
            sys.exit(2)
    else:
        logger.info("Nenhum mapeamento a ser aplicado no banco de dados.")

    if not ticket_ids:
        logger.info("Nenhum ticket encontrado no CSV para download. Encerrando.")
    else:
        base_dir = Path(args.download_dir)
        ensure_dir(base_dir)
        state_dir = base_dir / ".state"
        ensure_dir(state_dir)

        processed, saved = process_tickets(
            ticket_ids, cfg, args.download_dir, args.min_attach_kb,
            args.batch_size, state_dir, not args.no_attach_signature_block, logger
        )
        logger.info("Processamento de tickets concluído. Tickets processados: %d | Anexos salvos: %d", processed, saved)

    # >>> INÍCIO MOVER CSV E LOG <<<
    # Garante que os manipuladores de log de arquivo sejam fechados antes de mover o arquivo
    for handler in logger.handlers[:]:
        if isinstance(handler, logging.FileHandler):
            handler.close()
            logger.removeHandler(handler)

    # Move arquivos de log e de erros para a pasta 'old/'
    logger.info("Movendo arquivos de log e de erros para a pasta 'old/'.")
    try:
        old_dir = Path("old")
        ensure_dir(old_dir)

        # Move arquivo de log, se existir
        if args.log_file and Path(args.log_file).exists():
            shutil.move(args.log_file, old_dir / Path(args.log_file).name)
            logger.info("Arquivo '%s' movido para 'old/'.", args.log_file)

        # Move arquivo de erros, se existir
        if args.errors_file and Path(args.errors_file).exists():
            shutil.move(args.errors_file, old_dir / Path(args.errors_file).name)
            logger.info("Arquivo '%s' movido para 'old/'.", args.errors_file)

    except Exception as e:
        logger.error("Erro ao mover arquivos para a pasta 'old/': %s", e)
    # >>> FIM  <<<

if __name__ == "__main__":
    main()